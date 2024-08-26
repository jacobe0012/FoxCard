from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import exceltool
import os

#gen.bat需配置luban.gen_code_json.bat路径
# 输出的处理过的excel文件夹路径
lubanOutputPath = "D:\ConfigTools\Luban\ConfigRoot\Datas"
my_excel = []
folder_paths = []
def get_excel_files(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
                excel_files.append(file_path)
    return excel_files



# /// <summary>
# /// 在这里加入需要转换未处理的excel文件，包含文件后缀
# /// </summary>
my_excel.append(r"D:\JiYu\config\language.xlsx")
my_excel.append(r"D:\JiYu\config\art.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\area.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\obstacle.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\refresh.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\level.xlsx")

# /// <summary>
# /// 在这里加入需要转换未处理的excel整个文件夹，不包含文件后缀
# /// </summary>
folder_paths.append(r"D:\JiYu\config\activity")
folder_paths.append(r"D:\JiYu\config\battle")
folder_paths.append(r"D:\JiYu\config\common")
folder_paths.append(r"D:\JiYu\config\item")
folder_paths.append(r"D:\JiYu\config\task")








# 获取所有文件夹下的 XLSX 和 XLS 文件路径
excel_files = []
for folder_path in folder_paths:
    excel_files.extend(get_excel_files(folder_path))

my_excel.extend(excel_files)

tablespath = lubanOutputPath + "\__tables__.xlsx"
workbook = load_workbook(tablespath)  # 替换为实际的文件名
worksheet = workbook['Sheet1']  # 替换为实际的工作表名
worksheet.delete_rows(2, worksheet.max_row)
workbook.save(tablespath)
workbook.close()

for index, value in enumerate(my_excel):
    exceltool.genfixedexcel(index, value,lubanOutputPath)
    print(f"已转换第 {index+1} 个表格, 路径为: {value}")


    
print('done!')













