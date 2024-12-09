from openpyxl import Workbook, load_workbook
import os
import re
import shutil

# 打开Excel文件
def genfixedexcel(index,excel_path,output_path):
    pattern = r"\~\$"
    pattern2 = r"副本"
    if re.search(pattern, excel_path):
        return None
    if re.search(pattern2, excel_path):
        return None
    script_path = os.path.abspath(__file__)
    dir_path = os.path.dirname(script_path)
    parent_dir = os.path.dirname(dir_path)

    index = index * 10 +4
    folder_name = os.path.basename(os.path.dirname(excel_path))
    file_name = os.path.basename(excel_path)
    
        # 指定原始文件路径和目标文件路径
    file_outdir = output_path + "/"  + file_name
    olddir= parent_dir + r"\config\battle\skill_effect.xlsx"  
    # 复制原始文件到目标文件路径
    shutil.copy(excel_path, file_outdir)

    wb = load_workbook(file_outdir, data_only=True)
    # 选择第一个工作表
    #ws = wb['_data']
    sheet_names = wb.sheetnames
    
    #sheet_name = wb.sheetnames[0]
        #ws =wb[matching_sheets[0]] 
        # 删除除目标表之外的其他表
    for sheet_name in sheet_names:
        if not( '|' or "_self" in sheet_name):
            sheet = wb[sheet_name]
            wb.remove(sheet)     
        """    
        if "task|" in sheet_name:
            sheet = wb[sheet_name]
            new_ws =wb.create_sheet(title=sheet_name+r"New")
            for row in sheet.iter_rows(values_only=True):
                new_ws.append(row)

            wb.remove(sheet)
        """        
    sheet_names = wb.sheetnames
    # 遍历工作表名称，找到包含"data"的工作表
    matching_sheets = [sheet_name for sheet_name in sheet_names if '|' in sheet_name.lower()]

    matching_sheetself = [sheet_name for sheet_name in sheet_names if '_self' in sheet_name.lower()]

    # 处理自用的excel表
    if len(matching_sheetself) > 0:
        
        # 保存修改后的Excel文件
        file_name = os.path.basename(excel_path)
        file_name_without_extension = os.path.splitext(file_name)[0]
        new_file_name_without_extension = file_name_without_extension + "_Fixed.xlsx"
        myname = output_path + "/"  + file_name_without_extension + "_Fixed.xlsx"

        # 将此次处理的excel文件写入luban的__tables__.xlsx文件中
        
        tablespath = output_path + r"\__tables__.xlsx"
        workbook = load_workbook(tablespath)  # 替换为实际的文件名
        worksheet = workbook['Sheet1']  # 替换为实际的工作表名

        # TODO这里可能有单例表 映射表 列表 默认映射表(键为第一个)
        cell0 = worksheet.cell(row=index, column=2)
        cell0.value ="config.Tb" + file_name_without_extension   
                        
        cell1 = worksheet.cell(row=index, column=3)
        cell1.value = file_name_without_extension  
                                
        cell2 = worksheet.cell(row=index, column=4)
        cell2.value = "TRUE"
                            
        cell3 = worksheet.cell(row=index, column=5)
        cell3.value = new_file_name_without_extension
       
        # 保存luban.__tables__和兼容后的配置表
        workbook.save(tablespath)
        workbook.close()
        
        sheet = wb.active

        # 选择要排序的列，这里选择第5列（E列）对应的列索引为4
        column_index = 4
        column_index1 = 5
        # 从第二行开始获取数据
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        # 根据选中列的值对数据进行排序
        sorted_data = sorted(data, key=lambda x: x[column_index] if x[column_index] is not None else float('inf'))
        sorted_data = sorted(sorted_data, key=lambda x: x[column_index1] if x[column_index1] is not None else float('inf'))
        
        # 更新工作表中的数据，按照升序排序后的数据顺序进行排列
        for i, row_data in enumerate(sorted_data, start=2):
            for j, value in enumerate(row_data, start=1):
                sheet.cell(row=i, column=j, value=value)

        wb.save(myname)
        wb.close()
        return None

    if len(matching_sheets) == 0:
        print(excel_path + r"：找不到含有'|'的工作表")
        wb.close()
        return None
    
    tablespath = output_path + r"\__tables__.xlsx"
    workbook = load_workbook(tablespath)  # 替换为实际的文件名
    worksheet = workbook['Sheet1']  # 替换为实际的工作表名

    for idx, sheet in enumerate(matching_sheets, start=1): 
        resultName = sheet.split("|")[0]
        ws =wb[sheet] 
        
        if resultName == "event":
            resultName = "event_0"
        #if "task|" in sheet:
        #    continue
                      
        ws.title =resultName
        # 保存修改后的 Excel 文件
        #wb.save('updated_excel_file.xlsx')

        #ws = wb[sheet_name]

        # 插入空白列

        ws.insert_cols(1)

        # 向A1单元格写入数据
        ws['A1'] = '##var'
        ws['A2'] = '##type'
        ws['A3'] = '##group'
        ws['A4'] = '##'

        #给多语言表新增一个当前语言字段 current
        if resultName == r"language":
            last_column = ws.max_column +1  
            different_strings = ["current", "string", "client", "当前语言"]  
            for idx, string in enumerate(different_strings, start=1):  # Start from row 2
                cell = ws.cell(row=idx, column=last_column)  # Select the cell in the last column and current row
                cell.value = string
        
        if resultName == r"monster_template":
            last_column = ws.max_column +1     
            different_strings = ["monster_id", "int", "client", "怪物id"]  
            for idx, string in enumerate(different_strings, start=1):  # Start from row 2
                cell = ws.cell(row=idx, column=last_column)  # Select the cell in the last column and current row
                cell.value = string        
        


        # 修改第一行避免关键字冲突
        row1 = ws[1]
        for cell in row1[1:]:
            # 将单元格的值转换为全小写
            cell.value = str(cell.value).lower()
            #cell.value = str(cell.value).replace(" ", "")
            # 如果与对比字符串相等，则替换为新字符串
            if cell.value == "base":
                cell.value = "base_0"
            if cell.value == "event":
                cell.value = "event_0"
            if cell.value == "default":
                cell.value = "default_0"
            if cell.value == "goto":
                cell.value = "goto_0"    
            if cell.value == "lock":
                cell.value = "lock_0"

        # 获取总行数
        total_rows = ws.max_row
        # 创建一个列表来存储需要删除的行号
        rows_to_delete = []
        # 遍历每一行，检查第一列是否为空
        for row in range(1, total_rows + 1):
            cell_value = ws.cell(row=row, column=2).value
            if cell_value is None or cell_value == "":
                rows_to_delete.append(row)
        # 倒序删除空行，以避免删除后索引混乱

        for row in reversed(rows_to_delete):
            ws.delete_rows(row)

        # 修改第三行读取权限为luban兼容
        row3 = ws[3]
        for cell in row3[1:]:
            # 将单元格的值转换为全小写
            cell.value = str(cell.value).lower()
            #cell.value = str(cell.value).replace(" ", "")
            # 如果与对比字符串相等，则替换为新字符串
            if cell.value == "client":
                cell.value = "c"
            elif cell.value == "all":
                cell.value = "c,s"
            elif cell.value == "server": 
                cell.value = "s"
            else:
                cell.value = "none"        

        # 修改第二行数据类型为luban兼容
        row2 = ws[2]
        for cell in row2[1:]:
        # 将单元格的值转换为全小写
            cell.value = str(cell.value).lower()
            #cell.value = str(cell.value).replace(" ", "")
            # 如果与对比字符串相等，则替换为新字符串
            if cell.value == "array_string":
                cell.value = "(list#sep=;),string"
            elif cell.value == "array1_int" or cell.value == "array_int":
                cell.value = "(list#sep=;),int"
            elif cell.value == "array2_int":
                cell.value = "(list#sep=|),vector2"            
            elif cell.value == "array3_int":
                cell.value = "(list#sep=|),vector3"

            
            elif cell.value == "none":
                ws.delete_cols(cell.column)         
            # TODO:补充类型处
        # 获取第二行的数据


        # 保存修改后的Excel文件
        
        #file_name_without_extension = os.path.splitext(file_name)[0]
        
       

        
        new_file_name_without_extension = resultName + "_Fixed.xlsx"
        myname = output_path + "/"  + resultName + "_Fixed.xlsx"
        
        

        
        
        
        # 将此次处理的excel文件写入luban的__tables__.xlsx文件中
        


    # TODO这里可能有单例表 映射表 列表 默认映射表(键为第一个)
        cell0 = worksheet.cell(row=index, column=2)
        cell0.value ="config.Tb" + resultName   
                        
        cell1 = worksheet.cell(row=index, column=3)
        cell1.value = resultName  
                                
        cell2 = worksheet.cell(row=index, column=4)
        cell2.value = "TRUE"
                            
        cell3 = worksheet.cell(row=index, column=5)
        cell3.value =resultName + r"@" + file_name

        # 将equip_data表转为多键映射表

        if resultName ==r"equip_data":
            cell4 = worksheet.cell(row=index, column=6)
            cell4.value = "id+quality"

        # 将若干表转为单列表表
        listHandleArr = []
        # 在这里添加List的表路径

        listHandleArr.append(r"battle_drop")
        listHandleArr.append(r"monster_template")
        listHandleArr.append(r"drop")
        listHandleArr.append(r"draw_banner")
        listHandleArr.append(r"battleshop_drop")
        listHandleArr.append(r"monster_trigger")
        listHandleArr.append(r"battlepass_reward")
        listHandleArr.append(r"element_effect")
        listHandleArr.append(r"module_template")
        listHandleArr.append(r"battletech_drop")
        listHandleArr.append(r"fund_reward")
        listHandleArr.append(r"monopoly_shop")
        listHandleArr.append(r"turntable_level")     
        listHandleArr.append(r"turntable_score")  
        if resultName in listHandleArr:
            handleList(ws,worksheet,index)
        # 保存luban.__tables__和兼容后的配置表
        index +=1
        
        print(f"已转换表格: {resultName}")
    

    wb.save(file_outdir)    
    wb.close()     

    workbook.save(tablespath)    
    workbook.close()    

def handleList(ws,worksheet,index):
    cell4 = worksheet.cell(row=index, column=7)
    cell4.value = "list"
        # 计算实际的行数
    actual_max_row = ws.calculate_dimension().split(':')[1]
    # 从实际行数字符串中提取行号
    total_rows = int(actual_max_row[1:])
        # 在第二列之后插入一个新列
    ws.insert_cols(2)
    # 将值写入第二列的1-4行
    values = ['assist', 'int', 'c,s','辅助主键id']
    for i, value in enumerate(values, start=1):
        ws.cell(row=i, column=2, value=value)        
    # 从第5行开始，按从1到n的值填充新列
    start_row = 5
    n = total_rows - start_row + 1
    for i in range(1, n + 1):
        ws.cell(row=start_row + i - 1, column=2, value=i)        



    
