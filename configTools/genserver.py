from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import exceltoolserver
import os
script_path = os.path.abspath(__file__)
dir_path = os.path.dirname(script_path)
parent_dir = os.path.dirname(dir_path)

#gen.bat需配置luban.gen_code_json.bat路径
# 输出的处理过的excel文件夹路径
lubanOutputPath = dir_path + r"\Luban\ConfigRoot\Datas"
my_excel = []
folder_paths = []
def get_excel_files(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
                excel_files.append(file_path)
    return excel_files



# /// <summary>
# /// 在这里加入需要转换未处理的excel文件，包含文件后缀
# /// </summary>
#my_excel.append(r"D:\FoxCard\config\language.xlsx")

#my_excel.append(r"D:\FoxCard\config\foxconfig\card_group.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\area.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\obstacle.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\refresh.xlsx")
#my_excel.append(r"D:\JiYu\config\battle\level.xlsx")

# /// <summary>
# /// 在这里加入需要转换未处理的excel整个文件夹，不包含文件后缀
# /// </summary>
#folder_paths.append(r"D:\FoxCard\config\foxconfig")
#folder_paths.append(r"D:\JiYu\config\battle")
#folder_paths.append(r"D:\JiYu\config\common")
#folder_paths.append(r"D:\JiYu\config\item")
#folder_paths.append(r"D:\JiYu\config\task")
configDir =parent_dir + r"\config"
folder_paths.append(configDir)





# 获取所有文件夹下的 XLSX 和 XLS 文件路径
excel_files = []
for folder_path in folder_paths:
    excel_files.extend(get_excel_files(folder_path))

my_excel.extend(excel_files)

tablespath = lubanOutputPath + r"\__tables__.xlsx"
workbook = load_workbook(tablespath)  # 替换为实际的文件名
worksheet = workbook['Sheet1']  # 替换为实际的工作表名
worksheet.delete_rows(2, worksheet.max_row)
workbook.save(tablespath)
workbook.close()
#print('begin')
for index, value in enumerate(my_excel):
    exceltoolserver.genfixedexcel(index, value,lubanOutputPath)
    print(f"已转换第 {index+1} 个表格, 路径为: {value}")


    
print('done!')













